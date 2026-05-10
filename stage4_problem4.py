# file: stage4_problem4.py


import matplotlib

matplotlib.use("Agg")
import config  # 触发 config.py 中的字体配置

import time as _time
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from config import TaskConfig, data_path, plot_config, TABLE_DIR, PLOT_DIR, INTERMEDIATE_DIR, ensure_dirs
from core.constraint_checker import ConstraintChecker
from core.motion_utils import (
    compute_acceleration,
    compute_velocity,
)

try:
    plt.style.use("seaborn-v0_8-whitegrid")
except OSError:
    try:
        plt.style.use("seaborn-whitegrid")
    except OSError:
        pass

plot_config.apply_style()

try:
    import pulp

    HAS_PULP = True
except ImportError:
    HAS_PULP = False

try:
    from scipy.signal import savgol_filter

    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False

_COLOR_SHOOT = "#DC2626"
_COLOR_PHOTO = "#2563EB"
_COLOR_TRAJ = "#16A34A"


def load_trajectory() -> tuple:
    traj_path = Path(TABLE_DIR) / "Problem3_10Hz.xlsx"

    if not traj_path.exists():
        for ext in (".xlsx", ".xls", ".csv"):
            alt = traj_path.with_suffix(ext)
            if alt.exists():
                traj_path = alt
                break

    print(f"[问题4] 加载轨迹：{traj_path}")

    df = pd.read_excel(traj_path, engine="openpyxl")

    col_map = {
        "Time(s)": "t", "时间(s)": "t", "时间": "t", "t": "t",
        "X(m)": "x", "X坐标(m)": "x", "X坐标": "x", "x": "x",
        "Y(m)": "y", "Y坐标(m)": "y", "Y坐标": "y", "y": "y",
    }
    df = df.rename(columns=col_map)
    df = df[["t", "x", "y"]].apply(pd.to_numeric, errors="coerce").dropna()

    t = df["t"].values.astype(np.float64)
    x = df["x"].values.astype(np.float64)
    y = df["y"].values.astype(np.float64)

    print(f"[问题4] 轨迹：{len(t)} 个采样点，"
          f"[{t[0]:.2f}, {t[-1]:.2f}] s，dt={t[1] - t[0]:.4f} s")

    return t, x, y


def load_targets() -> list:
    file_path = Path(data_path.file4)

    if not file_path.exists():
        search_candidates = [
            file_path,
            Path("data") / file_path.name,
            Path("data") / file_path,
        ]
        found = False
        for candidate in search_candidates:
            if candidate.exists():
                file_path = candidate
                found = True
                break
            for ext in (".xlsx", ".xls", ".csv"):
                alt = candidate.with_suffix(ext)
                if alt.exists():
                    file_path = alt
                    found = True
                    break
            if found:
                break

    if not file_path.exists():
        raise FileNotFoundError(
            f"[问题4] 找不到目标文件，已尝试以下路径：\n"
            f"  - {data_path.file4}\n"
            f"  - data/{Path(data_path.file4).name}\n"
            f"  请检查 config.py 中 file4 的路径配置。"
        )

    print(f"[问题4] 加载目标：{file_path}")

    df = pd.read_excel(file_path, engine="openpyxl")

    col_map = {
        "目标编号": "name", "编号": "name", "编号/名称": "name",
        "类型": "type", "任务类型": "type",
        "X坐标(m)": "x", "X坐标": "x", "X": "x", "x": "x",
        "Y坐标(m)": "y", "Y坐标": "y", "Y": "y", "y": "y",
    }
    df = df.rename(columns=col_map)

    for col in ("x", "y"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df.dropna(subset=["x", "y"], inplace=True)
    df.reset_index(drop=True, inplace=True)

    all_targets = []
    for idx, row in df.iterrows():
        name = str(row.get("name", f"T{idx + 1}"))
        all_targets.append({
            "id": idx + 1,
            "name": name,
            "x": float(row["x"]),
            "y": float(row["y"]),
        })

    print(f"[问题4] 目标总数：{len(all_targets)} 个")
    for t in all_targets:
        print(f"    {t['name']:>6s}  ({t['x']:.1f}, {t['y']:.1f})")

    return all_targets


def sparse_sample_windows(windows: list, interval: float = 0.5) -> list:
    if not windows:
        return []

    groups = {}
    for w in windows:
        key = (w["target_id"], w.get("task_type", "shoot"))
        groups.setdefault(key, []).append(w)

    sampled = []
    for key, group in groups.items():
        group.sort(key=lambda w: w["t_exec"])
        last_t = -1e9
        for w in group:
            if w["t_exec"] - last_t >= interval:
                sampled.append(w)
                last_t = w["t_exec"]

    return sampled


def compress_windows(windows: list, dt: float) -> list:
    if not windows:
        return []

    groups = {}
    for w in windows:
        key = (w["target_id"], w.get("task_type", "shoot"))
        groups.setdefault(key, []).append(w)

    compressed = []
    for key, group in groups.items():
        group.sort(key=lambda w: w["t_exec"])

        seg_start = group[0]["t_exec"]
        seg_end = group[0]["t_exec"]
        seg_best = group[0]

        for w in group[1:]:
            if w["t_exec"] - seg_end <= 1.5 * dt:
                seg_end = w["t_exec"]
                if w.get("speed", 1e9) < seg_best.get("speed", 1e9):
                    seg_best = w
            else:
                rep = dict(seg_best)
                rep["t_exec"] = (seg_start + seg_end) / 2.0
                rep["t_start"] = rep["t_exec"] - 1.5
                compressed.append(rep)

                seg_start = w["t_exec"]
                seg_end = w["t_exec"]
                seg_best = w

        rep = dict(seg_best)
        rep["t_exec"] = (seg_start + seg_end) / 2.0
        rep["t_start"] = rep["t_exec"] - 1.5
        compressed.append(rep)

    return compressed


def schedule_ilp(
        windows_shoot: list,
        windows_photo: list,
        all_targets: list,
        min_task_sep: float = 0.2,  # [C2] 0.5 → 0.2
        max_shoot_per_target: int = 3,
        min_heading_diff: float = 15.0,  # [C4] 30° → 15°
        heading_bin_size: float = 15.0,  # [C4] 30° → 15°（bin 同步缩小）
        weight_unique_target: float = 1.0,  # [C6] 新：覆盖目标权重
        weight_photo: float = 0.1,  # [C6] 新：拍照权重
        weight_angle_diversity: float = 0.05,  # [C6] 新：角度多样性权重
        time_limit: int = 180,
) -> list:
    all_windows = []

    for w in windows_shoot:
        entry = dict(w)
        entry["task_type"] = "shoot"
        entry["global_idx"] = len(all_windows)
        all_windows.append(entry)

    for w in windows_photo:
        entry = dict(w)
        entry["task_type"] = "photo"
        entry["global_idx"] = len(all_windows)
        all_windows.append(entry)

    n_windows = len(all_windows)
    if n_windows == 0:
        return []

    num_bins = int(360 / heading_bin_size)

    for w in all_windows:
        if w["task_type"] == "photo":
            heading = w.get("heading", 0)
            w["bin_idx"] = int(((heading % 360) + 360) % 360 / heading_bin_size) % num_bins
        else:
            w["bin_idx"] = -1

    target_bin_set = set()
    for w in all_windows:
        if w["task_type"] == "photo":
            target_bin_set.add((w["target_id"], w["bin_idx"]))
    target_bin_list = sorted(target_bin_set)
    n_target_bins = len(target_bin_list)
    target_bin_to_idx = {tb: i for i, tb in enumerate(target_bin_list)}

    all_target_ids = set(w["target_id"] for w in all_windows)

    prob = pulp.LpProblem("TaskScheduling", pulp.LpMaximize)

    x = [pulp.LpVariable(f"x_{i}", cat="Binary") for i in range(n_windows)]
    y = [pulp.LpVariable(f"y_{j}", cat="Binary") for j in range(n_target_bins)]

    z = {tid: pulp.LpVariable(f"z_{tid}", cat="Binary") for tid in all_target_ids}

    objective = 0

    objective += weight_unique_target * pulp.lpSum(z.values())

    for i, w in enumerate(all_windows):
        if w["task_type"] == "photo":
            objective += weight_photo * x[i]

    for j in range(n_target_bins):
        objective += weight_angle_diversity * y[j]

    prob += objective

    windows_by_target = {}
    for i, w in enumerate(all_windows):
        windows_by_target.setdefault(w["target_id"], []).append(i)

    for tid, indices in windows_by_target.items():
        prob += (
            z[tid] <= pulp.lpSum(x[i] for i in indices),
            f"target_cover_{tid}",
        )

    print(f"  [ILP] 构建冲突矩阵（{n_windows} 个窗口）...")
    conflict_pairs = []

    exec_sorted = sorted(range(n_windows),
                         key=lambda i: all_windows[i]["t_exec"])

    for pos_i in range(n_windows):
        idx_i = exec_sorted[pos_i]
        t_i = all_windows[idx_i]["t_exec"]
        for pos_j in range(pos_i + 1, n_windows):
            idx_j = exec_sorted[pos_j]
            t_j = all_windows[idx_j]["t_exec"]
            if t_j - t_i >= min_task_sep:
                break
            conflict_pairs.append((idx_i, idx_j))

    print(f"  [ILP] 冲突对数量：{len(conflict_pairs)}")

    for (i, j) in conflict_pairs:
        prob += x[i] + x[j] <= 1, f"conflict_{i}_{j}"

    shoot_by_target = {}
    for i, w in enumerate(all_windows):
        if w["task_type"] == "shoot":
            shoot_by_target.setdefault(w["target_id"], []).append(i)

    for tid, indices in shoot_by_target.items():
        prob += (
            pulp.lpSum(x[i] for i in indices) <= max_shoot_per_target,
            f"shoot_limit_{tid}",
        )

    photo_windows_by_tb = {}
    for i, w in enumerate(all_windows):
        if w["task_type"] == "photo":
            tb = (w["target_id"], w["bin_idx"])
            photo_windows_by_tb.setdefault(tb, []).append(i)

    for j, (tid, bidx) in enumerate(target_bin_list):
        tb = (tid, bidx)
        indices = photo_windows_by_tb.get(tb, [])
        if indices:
            prob += y[j] <= pulp.lpSum(x[i] for i in indices), \
                f"bin_link_{tid}_{bidx}"

    for tb, indices in photo_windows_by_tb.items():
        if len(indices) > 1:
            prob += pulp.lpSum(x[i] for i in indices) <= 1, \
                f"bin_one_{tb[0]}_{tb[1]}"

    photo_by_target = {}
    for i, w in enumerate(all_windows):
        if w["task_type"] == "photo":
            photo_by_target.setdefault(w["target_id"], []).append(i)

    angle_conflict_count = 0
    for tid, indices in photo_by_target.items():
        for ii in range(len(indices)):
            for jj in range(ii + 1, len(indices)):
                wi = all_windows[indices[ii]]
                wj = all_windows[indices[jj]]
                h_i = wi.get("heading", 0)
                h_j = wj.get("heading", 0)
                diff = abs(h_i - h_j)
                d = min(diff, 360.0 - diff)
                if d < min_heading_diff:
                    prob += x[indices[ii]] + x[indices[jj]] <= 1, \
                        f"angle_conflict_{indices[ii]}_{indices[jj]}"
                    angle_conflict_count += 1

    print(f"  [ILP] 角度冲突约束：{angle_conflict_count} 条")

    n_vars = n_windows + n_target_bins + len(z)
    n_consts = (len(conflict_pairs) + len(shoot_by_target)
                + n_target_bins + angle_conflict_count + len(z))
    print(f"\n  [ILP] 开始求解（变量={n_vars}, "
          f"约束≈{n_consts}, 时限={time_limit}s）...")

    solver = pulp.PULP_CBC_CMD(
        timeLimit=time_limit,
        msg=1,
        gapRel=0.01,
    )

    t_start_solve = _time.time()
    prob.solve(solver)
    t_solve = _time.time() - t_start_solve

    status = pulp.LpStatus[prob.status]
    obj_val = pulp.value(prob.objective) if prob.objective else 0

    print(f"  [ILP] 求解完成：status={status}, "
          f"objective={obj_val:.2f}, 耗时={t_solve:.1f}s")

    scheduled_tasks = []

    if status in ("Optimal", "Feasible"):
        selected_indices = [
            i for i in range(n_windows) if pulp.value(x[i]) > 0.5
        ]

        for i in selected_indices:
            w = all_windows[i]
            scheduled_tasks.append({
                "target_id": w["target_id"],
                "task_type": w["task_type"],
                "t_start_prep": w["t_start"],
                "t_execute": w["t_exec"],
                "heading": w.get("heading"),
                "distance": w.get("distance"),
                "speed": w.get("speed"),
            })

        scheduled_tasks.sort(key=lambda t: t["t_execute"])

        covered_targets = sum(
            1 for tid in all_target_ids
            if pulp.value(z[tid]) > 0.5
        )
        covered_bins = sum(
            1 for j in range(n_target_bins)
            if pulp.value(y[j]) > 0.5
        )

        print(f"  [ILP] 选中窗口：{len(selected_indices)} 个")
        print(f"  [ILP] 覆盖目标：{covered_targets}/{len(all_target_ids)} 个")
        print(f"  [ILP] 角度bin覆盖：{covered_bins}/{n_target_bins} 个")
    else:
        print(f"  [ILP] 求解未成功（status={status}），将回退到贪心调度")

    return scheduled_tasks


def schedule_greedy_v2(
        windows_shoot: list,
        windows_photo: list,
        all_targets: list,
) -> list:
    MIN_TASK_SEP = 0.2  # [C2]
    MAX_SHOOT_PER_TARGET = 3
    MIN_HEADING_DIFF = 15.0  # [C4]
    HEADING_BIN_SIZE = 15.0  # [C4] 与 MIN_HEADING_DIFF 对齐
    NUM_HEADING_BINS = int(360 / HEADING_BIN_SIZE)

    id_to_name = {t["id"]: t["name"] for t in all_targets}
    scheduled_tasks = []

    shoot_windows_by_target = {}
    for w in windows_shoot:
        shoot_windows_by_target.setdefault(w["target_id"], []).append(w)

    mrv_shoot_targets = sorted(
        shoot_windows_by_target.keys(),
        key=lambda tid: len(shoot_windows_by_target[tid]),
    )

    shoot_count_map = {}

    for tid in mrv_shoot_targets:
        target_windows = sorted(
            shoot_windows_by_target[tid], key=lambda w: w["t_exec"]
        )
        for w in target_windows:
            if shoot_count_map.get(tid, 0) >= MAX_SHOOT_PER_TARGET:
                break

            time_conflict = False
            for st in scheduled_tasks:
                if abs(w["t_exec"] - st["t_execute"]) < MIN_TASK_SEP:
                    time_conflict = True
                    break
            if time_conflict:
                continue

            scheduled_tasks.append({
                "target_id": w["target_id"],
                "task_type": "shoot",
                "t_start_prep": w["t_start"],
                "t_execute": w["t_exec"],
                "heading": w.get("heading"),
                "distance": w.get("distance"),
                "speed": w.get("speed"),
            })
            shoot_count_map[tid] = shoot_count_map.get(tid, 0) + 1

    scheduled_tasks.sort(key=lambda tsk: tsk["t_execute"])

    shoot_count_total = len(scheduled_tasks)
    shoot_target_count = len({t["target_id"] for t in scheduled_tasks})
    print(f"\n  射击调度完成（MRV）：{shoot_count_total} 次射击, "
          f"覆盖 {shoot_target_count} 个目标")

    photo_windows_by_target = {}
    for w in windows_photo:
        photo_windows_by_target.setdefault(w["target_id"], []).append(w)

    mrv_photo_targets = sorted(
        photo_windows_by_target.keys(),
        key=lambda tid: len(photo_windows_by_target[tid]),
    )

    target_photo_bins = {}
    photo_count_total = 0

    for tid in mrv_photo_targets:
        target_windows = sorted(
            photo_windows_by_target[tid], key=lambda w: w["t_exec"]
        )
        for w in target_windows:
            heading = w.get("heading", 0)
            bin_idx = int(((heading % 360) + 360) % 360 / HEADING_BIN_SIZE) % NUM_HEADING_BINS

            if bin_idx in target_photo_bins.get(tid, set()):
                continue

            time_conflict = False
            for st in scheduled_tasks:
                if abs(w["t_exec"] - st["t_execute"]) < MIN_TASK_SEP:
                    time_conflict = True
                    break
            if time_conflict:
                continue

            angle_conflict = False
            existing_headings = [
                st.get("heading", 0)
                for st in scheduled_tasks
                if st["target_id"] == tid and st["task_type"] == "photo"
            ]
            for eh in existing_headings:
                diff = abs(heading - eh)
                d = min(diff, 360.0 - diff)
                if d < MIN_HEADING_DIFF:
                    angle_conflict = True
                    break
            if angle_conflict:
                continue

            scheduled_tasks.append({
                "target_id": w["target_id"],
                "task_type": "photo",
                "t_start_prep": w["t_start"],
                "t_execute": w["t_exec"],
                "heading": w.get("heading"),
                "distance": w.get("distance"),
                "speed": w.get("speed"),
            })

            if tid not in target_photo_bins:
                target_photo_bins[tid] = set()
            target_photo_bins[tid].add(bin_idx)
            photo_count_total += 1

    scheduled_tasks.sort(key=lambda tsk: tsk["t_execute"])

    photo_target_count = len(
        {t["target_id"] for t in scheduled_tasks if t["task_type"] == "photo"}
    )
    print(f"\n  拍照调度完成（MRV）：{photo_count_total} 次拍照, "
          f"覆盖 {photo_target_count} 个目标")

    shoot_count_map = {}
    for tsk in scheduled_tasks:
        if tsk["task_type"] == "shoot":
            tid = tsk["target_id"]
            shoot_count_map[tid] = shoot_count_map.get(tid, 0) + 1

    backfill_added = 0

    scheduled_tids_shoot = {
        tsk["target_id"] for tsk in scheduled_tasks
        if tsk["task_type"] == "shoot"
    }
    uncovered_tids = [
        tid for tid in set(w["target_id"] for w in windows_shoot)
        if tid not in scheduled_tids_shoot
    ]

    if uncovered_tids:
        print(f"  [调试] 未覆盖射击目标："
              f"{[id_to_name.get(tid, f'T{tid}') for tid in uncovered_tids]}")

        backfill_shoot_windows = []
        for w in windows_shoot:
            if w["target_id"] in uncovered_tids:
                backfill_shoot_windows.append(w)
        backfill_shoot_windows.sort(key=lambda w: w["t_exec"])

        for w in backfill_shoot_windows:
            tid = w["target_id"]
            if shoot_count_map.get(tid, 0) >= MAX_SHOOT_PER_TARGET:
                continue

            time_conflict = False
            for st in scheduled_tasks:
                if abs(w["t_exec"] - st["t_execute"]) < MIN_TASK_SEP:
                    time_conflict = True
                    break
            if time_conflict:
                continue

            scheduled_tasks.append({
                "target_id": w["target_id"],
                "task_type": "shoot",
                "t_start_prep": w["t_start"],
                "t_execute": w["t_exec"],
                "heading": w.get("heading"),
                "distance": w.get("distance"),
                "speed": w.get("speed"),
            })
            shoot_count_map[tid] = shoot_count_map.get(tid, 0) + 1
            backfill_added += 1
            name = id_to_name.get(tid, f"T{tid}")
            print(f"    回填射击：{name} @ {w['t_exec']:.2f}s")

    photo_windows_remaining = sorted(
        [w for w in windows_photo
         if w["target_id"] not in target_photo_bins
         or len(target_photo_bins.get(w["target_id"], set())) < NUM_HEADING_BINS],
        key=lambda w: w["t_exec"],
    )

    for w in photo_windows_remaining:
        tid = w["target_id"]
        heading = w.get("heading", 0)
        bin_idx = int(((heading % 360) + 360) % 360 / HEADING_BIN_SIZE) % NUM_HEADING_BINS

        if bin_idx in target_photo_bins.get(tid, set()):
            continue

        time_conflict = False
        for st in scheduled_tasks:
            if abs(w["t_exec"] - st["t_execute"]) < MIN_TASK_SEP:
                time_conflict = True
                break
        if time_conflict:
            continue

        angle_conflict = False
        existing_headings = [
            st.get("heading", 0)
            for st in scheduled_tasks
            if st["target_id"] == tid and st["task_type"] == "photo"
        ]
        for eh in existing_headings:
            diff = abs(heading - eh)
            d = min(diff, 360.0 - diff)
            if d < MIN_HEADING_DIFF:
                angle_conflict = True
                break
        if angle_conflict:
            continue

        scheduled_tasks.append({
            "target_id": w["target_id"],
            "task_type": "photo",
            "t_start_prep": w["t_start"],
            "t_execute": w["t_exec"],
            "heading": w.get("heading"),
            "distance": w.get("distance"),
            "speed": w.get("speed"),
        })

        if tid not in target_photo_bins:
            target_photo_bins[tid] = set()
        target_photo_bins[tid].add(bin_idx)
        backfill_added += 1

    scheduled_tasks.sort(key=lambda tsk: tsk["t_execute"])
    print(f"\n  回填新增：{backfill_added} 个任务")

    return scheduled_tasks


def print_summary(
        scheduled_tasks: list,
        all_targets: list,
) -> None:
    shoot_tasks = [t for t in scheduled_tasks if t["task_type"] == "shoot"]
    photo_tasks = [t for t in scheduled_tasks if t["task_type"] == "photo"]

    print("\n" + "=" * 60)
    print("  问题 4 结果汇总")
    print("=" * 60)
    print(f"  总任务数：{len(scheduled_tasks)}")
    print(f"  射击任务：{len(shoot_tasks)} 个")
    print(f"  拍照任务：{len(photo_tasks)} 个")

    if len(scheduled_tasks) > 0:
        t_first = scheduled_tasks[0]["t_start_prep"]
        t_last = scheduled_tasks[-1]["t_execute"]
        print(f"  时间跨度：[{t_first:.2f}, {t_last:.2f}] s")
        print(f"  总耗时：{t_last - t_first:.2f} s")

    if len(photo_tasks) > 0:
        print("\n  拍照目标详情：")
        id_to_name = {t["id"]: t["name"] for t in all_targets}

        photo_by_target: dict = {}
        for pt in photo_tasks:
            tid = pt["target_id"]
            photo_by_target.setdefault(tid, [])
            photo_by_target[tid].append(pt.get("heading", None))

        for tid, headings in photo_by_target.items():
            name = id_to_name.get(tid, f"目标{tid}")
            heading_strs = []
            for h in headings:
                if h is not None:
                    heading_strs.append(f"{h:.1f} deg")
                else:
                    heading_strs.append("N/A")
            print(f"    {name}: {len(headings)} 次, "
                  f"角度=[{', '.join(heading_strs)}]")

    print("=" * 60)


def plot_problem4_results(
        t: np.ndarray, x: np.ndarray, y: np.ndarray,
        all_targets: list,
        scheduled_tasks: list,
        output_dir: Path,
) -> None:
    figures_dir = Path(PLOT_DIR)
    figures_dir.mkdir(parents=True, exist_ok=True)

    fig, ax = plt.subplots(figsize=(14, 10))

    ax.plot(x, y, color=_COLOR_TRAJ, linewidth=0.6, alpha=0.7,
            label="融合轨迹")

    for tgt in all_targets:
        ax.scatter(tgt["x"], tgt["y"], s=120, marker="^",
                   color=_COLOR_SHOOT, edgecolors="black",
                   linewidths=0.8, zorder=5)
        ax.annotate(tgt["name"], (tgt["x"], tgt["y"]),
                    textcoords="offset points", xytext=(8, 8),
                    fontsize=9, fontweight="bold", color=_COLOR_SHOOT)

    id_to_info = {tgt["id"]: tgt for tgt in all_targets}

    for task in scheduled_tasks:
        tid = task["target_id"]
        tgt = id_to_info.get(tid)
        if tgt is None:
            continue
        i_exec = np.argmin(np.abs(t - task["t_execute"]))
        color = _COLOR_SHOOT if task["task_type"] == "shoot" else _COLOR_PHOTO
        ax.plot([x[i_exec], tgt["x"]], [y[i_exec], tgt["y"]],
                color=color, linewidth=0.5, linestyle="--", alpha=0.5)

    ax.set_xlabel("X (m)", fontsize=12)
    ax.set_ylabel("Y (m)", fontsize=12)
    ax.set_title(
        f"问题4 — 任务调度（共{len(scheduled_tasks)}项任务）",
        fontsize=14, fontweight="bold",
    )
    ax.legend(loc="upper right", fontsize=10)
    ax.set_aspect("equal", adjustable="datalim")

    fig.tight_layout()
    fig.savefig(figures_dir / "Problem4_schedule.png", dpi=180,
                bbox_inches="tight")
    plt.close(fig)

    if len(scheduled_tasks) > 0:
        fig, ax = plt.subplots(
            figsize=(16, max(4, len(scheduled_tasks) * 0.5))
        )

        y_labels = []
        for i, task in enumerate(scheduled_tasks):
            tid = task["target_id"]
            tgt = id_to_info.get(tid)
            name = tgt["name"] if tgt else f"T{tid}"
            task_label = "射击" if task["task_type"] == "shoot" else "拍照"
            y_labels.append(f"{name} ({task_label})")

            color = (
                _COLOR_SHOOT if task["task_type"] == "shoot" else _COLOR_PHOTO
            )

            ax.barh(
                i, task["t_execute"] - task["t_start_prep"],
                left=task["t_start_prep"],
                height=0.6, color=color, alpha=0.3, edgecolor=color,
            )
            ax.plot(task["t_execute"], i, marker="|", markersize=15,
                    color=color, markeredgewidth=2)

        ax.set_yticks(range(len(y_labels)))
        ax.set_yticklabels(y_labels, fontsize=9, fontfamily="sans-serif")
        ax.set_xlabel("Time (s)", fontsize=12)
        ax.set_title(
            "任务调度 — 甘特图",
            fontsize=14, fontweight="bold",
        )
        ax.invert_yaxis()

        fig.tight_layout()
        fig.savefig(figures_dir / "Problem4_gantt.png", dpi=180,
                    bbox_inches="tight")
        plt.close(fig)

    print(f"[问题4] 图表已保存至 {figures_dir}")


def plot_window_heatmap(
        windows_shoot: list,
        windows_photo: list,
        all_targets: list,
        scheduled_tasks: list,
        output_dir: Path,
) -> None:
    from matplotlib.lines import Line2D

    figures_dir = Path(PLOT_DIR)
    figures_dir.mkdir(parents=True, exist_ok=True)

    id_to_name = {tgt["id"]: tgt["name"] for tgt in all_targets}
    target_names = [tgt["name"] for tgt in all_targets]
    target_ids = [tgt["id"] for tgt in all_targets]
    n_targets = len(all_targets)

    fig, ax = plt.subplots(figsize=(16, max(6, n_targets * 0.4)))

    for w in windows_shoot:
        tid = w["target_id"]
        if tid in target_ids:
            y_idx = target_ids.index(tid)
            ax.scatter(w["t_exec"], y_idx, s=8,
                       c=plot_config.COLORS[4], alpha=0.4,
                       marker="s", edgecolors="none")

    for w in windows_photo:
        tid = w["target_id"]
        if tid in target_ids:
            y_idx = target_ids.index(tid)
            ax.scatter(w["t_exec"], y_idx, s=8,
                       c=plot_config.COLORS[5], alpha=0.4,
                       marker="s", edgecolors="none")

    for task in scheduled_tasks:
        tid = task["target_id"]
        if tid not in target_ids:
            continue
        y_idx = target_ids.index(tid)
        if task["task_type"] == "shoot":
            ax.scatter(task["t_execute"], y_idx, s=80,
                       c=plot_config.COLORS[0], marker="*",
                       edgecolors="black", linewidths=0.5, zorder=5)
        else:
            ax.scatter(task["t_execute"], y_idx, s=50,
                       c=plot_config.COLORS[3], marker="D",
                       edgecolors="black", linewidths=0.5, zorder=5)

    ax.set_yticks(range(n_targets))
    ax.set_yticklabels(target_names, fontsize=plot_config.tick_fontsize)
    ax.set_xlabel("Time (s)", fontsize=plot_config.label_fontsize)
    ax.set_ylabel("Target", fontsize=plot_config.label_fontsize)
    ax.set_title("可行窗口时空分布与调度结果",
                 fontsize=plot_config.title_fontsize, fontweight="bold")
    ax.invert_yaxis()

    legend_elements = [
        Line2D([0], [0], marker="s", color="w",
               markerfacecolor=plot_config.COLORS[4],
               markersize=8, label="射击窗口"),
        Line2D([0], [0], marker="s", color="w",
               markerfacecolor=plot_config.COLORS[5],
               markersize=8, label="拍照窗口"),
        Line2D([0], [0], marker="*", color="w",
               markerfacecolor=plot_config.COLORS[0],
               markersize=12, markeredgecolor="black",
               label="已调度-射击"),
        Line2D([0], [0], marker="D", color="w",
               markerfacecolor=plot_config.COLORS[3],
               markersize=8, markeredgecolor="black",
               label="已调度-拍照"),
    ]
    ax.legend(handles=legend_elements, loc="upper right",
              fontsize=plot_config.legend_fontsize,
              frameon=plot_config.legend_frameon)

    fig.tight_layout()
    fig.savefig(figures_dir / "Problem4_window_heatmap.png", dpi=plot_config.dpi)
    plt.close(fig)
    print(f"[问题4] 窗口时空分布图已保存")


def plot_schedule_comparison(
        windows_shoot: list,
        windows_photo: list,
        scheduled_tasks: list,
        all_targets: list,
        output_dir: Path,
) -> None:
    figures_dir = Path(PLOT_DIR)
    figures_dir.mkdir(parents=True, exist_ok=True)

    target_names = [tgt["name"] for tgt in all_targets]
    target_ids = [tgt["id"] for tgt in all_targets]
    n_targets = len(all_targets)

    fig, (ax1, ax2) = plt.subplots(
        1, 2, figsize=(20, max(6, n_targets * 0.4)), sharey=True,
    )

    for w in windows_shoot:
        tid = w["target_id"]
        if tid in target_ids:
            y_idx = target_ids.index(tid)
            ax1.barh(y_idx, 0.8, left=w["t_exec"] - 0.4,
                     height=0.6, color=plot_config.COLORS[4],
                     alpha=0.3, edgecolor="none")

    for w in windows_photo:
        tid = w["target_id"]
        if tid in target_ids:
            y_idx = target_ids.index(tid)
            ax1.barh(y_idx, 0.8, left=w["t_exec"] - 0.4,
                     height=0.6, color=plot_config.COLORS[5],
                     alpha=0.3, edgecolor="none")

    ax1.set_yticks(range(n_targets))
    ax1.set_yticklabels(target_names, fontsize=plot_config.tick_fontsize)
    ax1.set_xlabel("Time (s)", fontsize=plot_config.label_fontsize)
    ax1.set_title(
        f"调度前 — 可行窗口（射击{len(windows_shoot)}+拍照{len(windows_photo)}）",
        fontsize=plot_config.title_fontsize,
    )
    ax1.invert_yaxis()

    for task in scheduled_tasks:
        tid = task["target_id"]
        if tid not in target_ids:
            continue
        y_idx = target_ids.index(tid)
        color = (plot_config.COLORS[0] if task["task_type"] == "shoot"
                 else plot_config.COLORS[3])
        ax2.barh(y_idx, task["t_execute"] - task["t_start_prep"],
                 left=task["t_start_prep"], height=0.6,
                 color=color, alpha=0.6, edgecolor=color, linewidth=1)
        ax2.plot(task["t_execute"], y_idx, marker="|", markersize=15,
                 color=color, markeredgewidth=2)

    ax2.set_xlabel("Time (s)", fontsize=plot_config.label_fontsize)
    ax2.set_title(
        f"调度后 — 最终任务（共{len(scheduled_tasks)}项）",
        fontsize=plot_config.title_fontsize,
    )

    fig.suptitle("调度前后对比",
                 fontsize=plot_config.title_fontsize + 2,
                 fontweight="bold", y=1.01)
    fig.tight_layout()
    fig.savefig(figures_dir / "Problem4_schedule_comparison.png",
                dpi=plot_config.dpi, bbox_inches="tight")
    plt.close(fig)
    print(f"[问题4] 调度前后对比图已保存")


if __name__ == "__main__":
    output_dir = data_path.output_dir
    ensure_dirs()

    print("[问题4] 数据加载")
    print("=" * 60)
    t, x, y = load_trajectory()

    all_targets = load_targets()

    print("\n" + "=" * 60)
    print("  步骤3：运动状态计算")
    print("=" * 60)

    vx, vy, speed = compute_velocity(t, x, y)
    ax, ay, acc = compute_acceleration(t, x, y)

    print(f"  速率：min={np.min(speed):.3f}，max={np.max(speed):.3f}，"
          f"mean={np.mean(speed):.3f} m/s")
    print(f"  加速度：min={np.min(acc):.3f}，max={np.max(acc):.3f}，"
          f"mean={np.mean(acc):.3f} m/s²")

    print(f"  [C5] 加速度由 compute_acceleration 内部 SG 滤波得到，无重复平滑")


    print("\n" + "=" * 60)
    print("  步骤4：约束检查器实例化")
    print("=" * 60)

    checker = ConstraintChecker(t, x, y, vx, vy, ax, ay)
    print(f"  dt = {checker.dt:.4f} s，{checker.n} 个采样点")
    print(f"  射击距离约束：[{TaskConfig.SHOOT_DIST_MIN}, "
          f"{TaskConfig.SHOOT_DIST_MAX}] m")
    print(f"  射击速率上限：{TaskConfig.SHOOT_SPEED_MAX} m/s")
    print(f"  拍照距离约束：[{TaskConfig.PHOTO_DIST_MIN}, "
          f"{TaskConfig.PHOTO_DIST_MAX}] m")
    print(f"  拍照速率上限：{TaskConfig.PHOTO_SPEED_MAX} m/s")

    print("\n" + "=" * 60)
    print("  步骤5：搜索可行时间窗口")
    print("=" * 60)

    windows_shoot = checker.find_all_feasible_windows(
        all_targets, task_type="shoot", step_time=0.5,
    )
    for w in windows_shoot:
        w["task_type"] = "shoot"

    windows_photo = checker.find_all_feasible_windows(
        all_targets, task_type="photo", step_time=0.5,
    )
    for w in windows_photo:
        w["task_type"] = "photo"

    n_shoot_before = len(windows_shoot)
    n_photo_before = len(windows_photo)
    windows_shoot = sparse_sample_windows(windows_shoot, interval=0.5)
    windows_photo = sparse_sample_windows(windows_photo, interval=0.5)
    print(f"  [C1] 稀疏采样：射击 {n_shoot_before}→{len(windows_shoot)}，"
          f"拍照 {n_photo_before}→{len(windows_photo)}")

    print(f"  射击可行窗口（0.5s采样）：{len(windows_shoot)} 个")
    print(f"  拍照可行窗口（0.5s采样）：{len(windows_photo)} 个")

    id_to_name = {tgt["id"]: tgt["name"] for tgt in all_targets}

    for label, windows in [("射击", windows_shoot), ("拍照", windows_photo)]:
        if len(windows) > 0:
            counts = {}
            for w in windows:
                tid = w["target_id"]
                counts[tid] = counts.get(tid, 0) + 1
            print(f"  {label}窗口分布:")
            for tid, cnt in sorted(counts.items()):
                print(f"    {id_to_name.get(tid, f'T{tid}'):>6s}: {cnt} 个窗口")

    windows_all = windows_shoot + windows_photo

    if len(windows_all) > 0:
        print("\n  窗口时间分布 (前 30 个):")
        for w in sorted(windows_all, key=lambda x: x["t_exec"])[:30]:
            name = id_to_name.get(w["target_id"], f"T{w['target_id']}")
            print(
                f"    {name:>6s} {w['task_type']:6s}  "
                f"t_start={w['t_start']:.2f}  t_exec={w['t_exec']:.2f}  "
                f"dist={w['distance']:.1f}m  speed={w['speed']:.2f}m/s"
            )

    if len(windows_all) == 0:
        print("\n[问题4] 警告：未找到任何可行窗口，无法调度。")
        print("[问题4] 可能原因：")
        print("  1. 目标坐标与轨迹坐标系不匹配（轨迹 X:[{:.0f},{:.0f}], "
              "目标 X:[{:.0f},{:.0f}]）".format(
            x.min(), x.max(),
            min(t["x"] for t in all_targets),
            max(t["x"] for t in all_targets),
        ))
        print("  2. 速率/加速度约束过严（轨迹平均速率 {:.2f} m/s）".format(
            np.mean(speed)))
        df_empty = pd.DataFrame(columns=[
            "序号", "目标编号", "任务", "开始准备时刻(s)", "任务执行时刻(s)"
        ])
        df_empty.to_excel(
            Path(TABLE_DIR) / "result.xlsx", index=False, engine="openpyxl"
        )
        print(f"[问题4] 空结果已保存至 {TABLE_DIR}/result.xlsx")
        exit(0)

    print("\n" + "=" * 60)
    print("  步骤6：任务调度")
    print("=" * 60)

    scheduled_tasks = []
    solver_used = "none"

    if HAS_PULP:
        print("\n  [调度] 使用 ILP 数学优化建模（pulp 可用）")
        print(f"  [调度] [C6] 目标函数：max(Σz_t + {0.1}·photo + {0.05}·angle)")
        print(f"  [调度] [C2] MIN_TASK_SEP = 0.2s")
        print(f"  [调度] [C4] MIN_HEADING_DIFF = 15°")

        scheduled_tasks = schedule_ilp(
            windows_shoot=windows_shoot,
            windows_photo=windows_photo,
            all_targets=all_targets,
            min_task_sep=0.2,  # [C2]
            max_shoot_per_target=3,
            min_heading_diff=15.0,  # [C4]
            heading_bin_size=15.0,  # [C4]
            weight_unique_target=1.0,  # [C6]
            weight_photo=0.1,  # [C6]
            weight_angle_diversity=0.05,  # [C6]
            time_limit=180,
        )

        if scheduled_tasks:
            solver_used = "ILP"
        else:
            print("\n  [调度] ILP 未返回可行解，回退到贪心调度")

    if not scheduled_tasks:
        if not HAS_PULP:
            print("\n  [调度] pulp 未安装，使用贪心调度（pip install pulp 可启用ILP）")
        else:
            print("\n  [调度] 回退到贪心调度")

        print("  [调度] 使用 v3 优化版贪心调度（MRV + 全局时间排序 + 角度区间覆盖）")
        scheduled_tasks = schedule_greedy_v2(
            windows_shoot=windows_shoot,
            windows_photo=windows_photo,
            all_targets=all_targets,
        )
        solver_used = solver_used if solver_used == "ILP" else "Greedy_v3"

    print(f"\n  [调度] 求解器：{solver_used}")
    print(f"  [调度] 调度完成：{len(scheduled_tasks)} 个任务")

    shoot_per_target = {}
    photo_per_target = {}
    for tsk in scheduled_tasks:
        tid = tsk["target_id"]
        if tsk["task_type"] == "shoot":
            shoot_per_target[tid] = shoot_per_target.get(tid, 0) + 1
        else:
            photo_per_target[tid] = photo_per_target.get(tid, 0) + 1

    all_target_ids = set(shoot_per_target.keys()) | set(photo_per_target.keys())
    all_target_ids |= {tgt["id"] for tgt in all_targets}

    print("\n  各目标调度结果：")
    for tid in sorted(all_target_ids):
        name = id_to_name.get(tid, f"T{tid}")
        sc = shoot_per_target.get(tid, 0)
        pc = photo_per_target.get(tid, 0)
        status = ""
        if sc == 0 and pc == 0:
            status = " ← 未覆盖"
        print(f"    {name}: 射击 {sc} 次, 拍照 {pc} 次{status}")

    if len(scheduled_tasks) >= 2:
        gaps = []
        for i in range(len(scheduled_tasks) - 1):
            t_end = scheduled_tasks[i]["t_execute"]
            t_next = scheduled_tasks[i + 1]["t_start_prep"]
            gap = t_next - t_end
            if gap > 10:
                gaps.append((gap, t_end, t_next))

        if gaps:
            gaps.sort(key=lambda g: g[0], reverse=True)
            max_gap, gap_start, gap_end = gaps[0]
            print(f"\n  最大空闲区间：{max_gap:.1f} s "
                  f"[{gap_start:.2f}, {gap_end:.2f}]")
            if len(gaps) > 1:
                print(f"  其他较大空闲（>10s）：")
                for gap, gs, ge in gaps[:5]:
                    print(f"    {gap:.1f} s [{gs:.2f}, {ge:.2f}]")
            if max_gap > 60:
                print("  警告：存在超过 60 秒的连续空闲区间")
        else:
            print("\n  无超过 10 秒的空闲区间")

    print("\n" + "=" * 60)
    print("  步骤7：生成结果表")
    print("=" * 60)

    rows = []
    for i, task in enumerate(scheduled_tasks, start=1):
        task_label = "射击" if task["task_type"] == "shoot" else "拍照"
        rows.append({
            "序号": i,
            "目标编号": id_to_name.get(
                task["target_id"], f"T{task['target_id']}"
            ),
            "任务": task_label,
            "开始准备时刻(s)": round(task["t_start_prep"], 2),
            "任务执行时刻(s)": round(task["t_execute"], 2),
        })

    df_result = pd.DataFrame(rows)
    df_result = df_result.sort_values("任务执行时刻(s)").reset_index(drop=True)
    df_result["序号"] = range(1, len(df_result) + 1)

    print(f"\n{df_result.to_string(index=False)}")

    xlsx_path = Path(TABLE_DIR) / "result.xlsx"

    template_path = Path(data_path.file4).parent / "result_template.xlsx"
    if template_path.exists():
        try:
            from openpyxl import load_workbook

            wb = load_workbook(template_path)
            ws = wb.active

            for r_idx, row in enumerate(rows, start=2):
                ws.cell(row=r_idx, column=1, value=row["序号"])
                ws.cell(row=r_idx, column=2, value=row["目标编号"])
                ws.cell(row=r_idx, column=3, value=row["任务"])
                ws.cell(row=r_idx, column=4, value=row["开始准备时刻(s)"])
                ws.cell(row=r_idx, column=5, value=row["任务执行时刻(s)"])

            wb.save(xlsx_path)
            print(f"[问题4] 结果已保存至 {xlsx_path}（模板格式）")
        except Exception as e:
            print(f"[问题4] 模板写入失败（{e}），使用默认格式")
            df_result.to_excel(xlsx_path, index=False, engine="openpyxl")
            print(f"[问题4] 结果已保存至 {xlsx_path}")
    else:
        df_result.to_excel(xlsx_path, index=False, engine="openpyxl")
        print(f"[问题4] 结果已保存至 {xlsx_path}")

    size_kb = xlsx_path.stat().st_size / 1024
    print(f"[问题4] 文件大小：{size_kb:.1f} KB")

    print_summary(scheduled_tasks, all_targets)

    plot_problem4_results(
        t, x, y,
        all_targets,
        scheduled_tasks,
        output_dir,
    )

    print("\n" + "=" * 60)
    print("  约束满足率统计")
    print("=" * 60)

    total_shoot = len(windows_shoot)
    total_photo = len(windows_photo)
    total_candidates = total_shoot + total_photo

    sched_shoot = len([t for t in scheduled_tasks if t["task_type"] == "shoot"])
    sched_photo = len([t for t in scheduled_tasks if t["task_type"] == "photo"])
    total_scheduled = len(scheduled_tasks)

    covered_ids = set(t["target_id"] for t in scheduled_tasks)
    uncovered = len(all_targets) - len(covered_ids)

    print(f"  候选窗口：{total_candidates}（射击{total_shoot} + 拍照{total_photo}）")
    print(f"  最终调度：{total_scheduled}（射击{sched_shoot} + 拍照{sched_photo}）")
    print(f"  目标覆盖：{len(covered_ids)}/{len(all_targets)} "
          f"({100 * len(covered_ids) / max(len(all_targets), 1):.1f}%)")
    print(f"  未覆盖目标：{uncovered} 个")

    df_cstat = pd.DataFrame([
        {"阶段": "候选窗口", "数量": total_candidates, "类别": "全部"},
        {"阶段": "候选窗口", "数量": total_shoot, "类别": "射击"},
        {"阶段": "候选窗口", "数量": total_photo, "类别": "拍照"},
        {"阶段": "最终调度", "数量": total_scheduled, "类别": "全部"},
        {"阶段": "最终调度", "数量": sched_shoot, "类别": "射击"},
        {"阶段": "最终调度", "数量": sched_photo, "类别": "拍照"},
        {"阶段": "未覆盖目标", "数量": uncovered, "类别": "全部"},
    ])
    df_cstat.to_excel(Path(TABLE_DIR) / "constraint_stats.xlsx",
                      index=False, engine="openpyxl")
    print(f"  约束统计数据已保存至 {TABLE_DIR}/constraint_stats.xlsx")

    plot_window_heatmap(windows_shoot, windows_photo,
                        all_targets, scheduled_tasks, output_dir)
    plot_schedule_comparison(windows_shoot, windows_photo,
                             scheduled_tasks, all_targets, output_dir)

    import pickle

    result = {
        "t_fused":          t,
        "traj_x":           x,
        "traj_y":           y,
        "vx":               vx,
        "vy":               vy,
        "speed":            speed,
        "ax":               ax,
        "ay":               ay,
        "acc":              acc,
        "tasks":            scheduled_tasks,
        "all_targets":      all_targets,
        "windows_shoot":    windows_shoot,
        "windows_photo":    windows_photo,
        "solver_used":      solver_used,
        "n_shoot":          len([t_ for t_ in scheduled_tasks if t_["task_type"] == "shoot"]),
        "n_photo":          len([t_ for t_ in scheduled_tasks if t_["task_type"] == "photo"]),
        "covered_targets":  sorted(set(t_["target_id"] for t_ in scheduled_tasks)),
    }

    pkl_path = Path(INTERMEDIATE_DIR) / "result_problem4.pkl"
    with open(pkl_path, "wb") as _f:
        pickle.dump(result, _f, protocol=pickle.HIGHEST_PROTOCOL)
    print(f"\n[问题4] 可视化数据已保存 → {pkl_path}")
    print(f"  字段: {list(result.keys())}")

    print(f"\n[问题4] 问题4求解完毕。（求解器：{solver_used}）")
