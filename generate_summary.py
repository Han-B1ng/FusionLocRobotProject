# file: generate_summary.py
# @Author : Han_B1ng
# @Time : 2026/5/9
# @Description : 汇总表生成模块 — 从各阶段输出文件提取关键结果，生成五个标准化汇总表

"""
╔══════════════════════════════════════════════════════╗
║  汇总表生成模块                                       ║
╚══════════════════════════════════════════════════════╝

从各阶段的输出文件（.xlsx）以及阶段运行日志中提取关键指标，
生成五张标准化汇总表（Excel多Sheet），供论文/报告直接使用。

使用方法：
  python generate_summary.py                        # 运行全部4阶段→生成汇总表
  python generate_summary.py --output my_tables.xlsx # 指定输出路径

表结构：
  Sheet 1 — 问题一关键结果汇总
  Sheet 2 — 问题二关键结果汇总
  Sheet 3 — 问题三关键结果汇总
  Sheet 4 — 问题四任务调度结果
  Sheet 5 — 跨问题指标对比

依赖：config.py, core.*, openpyxl, pandas, numpy
"""

from __future__ import annotations

import argparse
import sys
import warnings
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

# ── 环境准备 ──
_PROJECT_ROOT = Path(__file__).resolve().parent
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(_PROJECT_ROOT))

from config import (
    alignment_config,
    data_path,
    filter_config,
    time_config,
    TABLE_DIR,
    ensure_dirs,
    _detected_font,
)

# 使用系统检测到的中文字体（openpyxl 用）
_DETECTED_CJK = _detected_font or "SimHei"

# ============================================================
#  Excel 样式
# ============================================================
try:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False


def _default_styles():
    """返回默认样式字典，openpyxl 不可用时返回空字典。"""
    if not _HAS_OPENPYXL:
        return {}
    return {
        "title_font": Font(name="SimHei", size=13, bold=True),
        "header_font": Font(name="SimHei", size=11, bold=True, color="FFFFFF"),
        "cell_font": Font(name="SimHei", size=10.5),
        "header_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "thin_border": Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        ),
        "center_align": Alignment(horizontal="center", vertical="center"),
        "left_align": Alignment(horizontal="left", vertical="center", wrap_text=True),
    }


def _style_two_col_sheet(ws, start_row, n_data_rows):
    """给双列表格套用统一样式。"""
    if not _HAS_OPENPYXL:
        return
    s = _default_styles()
    for row in range(start_row, start_row + n_data_rows + 1):
        for col in (1, 2):
            cell = ws.cell(row=row, column=col)
            cell.border = s["thin_border"]
            if row == start_row:
                cell.font = s["header_font"]
                cell.fill = s["header_fill"]
                cell.alignment = s["center_align"]
            else:
                cell.font = s["cell_font"]
                cell.alignment = s["center_align"] if col == 2 else s["left_align"]


def _style_multi_col_sheet(ws, start_row, n_data_rows, n_cols):
    """给多列表格套用统一样式。"""
    if not _HAS_OPENPYXL:
        return
    s = _default_styles()
    for row in range(start_row, start_row + n_data_rows + 1):
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = s["thin_border"]
            if row == start_row:
                cell.font = s["header_font"]
                cell.fill = s["header_fill"]
                cell.alignment = s["center_align"]
            else:
                cell.font = s["cell_font"]
                cell.alignment = s["center_align"] if col > 1 else s["left_align"]


# ============================================================
#  各问题数据提取函数
# ============================================================

def _extract_problem1_data() -> Dict[str, str]:
    """从 Problem1_10Hz.xlsx 和 stage1 运行结果提取问题一关键指标。"""
    from core.time_alignment import align_sensors

    # 加载附件1原始数据
    file_path = data_path.path1
    df1 = pd.read_excel(file_path, sheet_name="方式1(4Hz)", engine="openpyxl")
    df2 = pd.read_excel(file_path, sheet_name="方式2(5Hz)", engine="openpyxl")

    col_map = {"时间(s)": "t", "X坐标(m)": "x", "Y坐标(m)": "y"}
    df1 = df1.rename(columns=col_map)[["t", "x", "y"]].apply(pd.to_numeric, errors="coerce").dropna()
    df2 = df2.rename(columns=col_map)[["t", "x", "y"]].apply(pd.to_numeric, errors="coerce").dropna()

    t1 = df1["t"].values.astype(np.float64)
    x1 = df1["x"].values.astype(np.float64)
    y1 = df1["y"].values.astype(np.float64)
    t2 = df2["t"].values.astype(np.float64)
    x2 = df2["x"].values.astype(np.float64)
    y2 = df2["y"].values.astype(np.float64)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        delay_fine, t_grid, _, _, _, _ = align_sensors(
            t1, x1, y1, t2, x2, y2,
            target_freq=time_config.target_freq,
            delay_range=alignment_config.delay_range,
            method=alignment_config.method,
            w1=0.5, w2=0.5,
        )

    direction = "传感器2滞后于传感器1" if delay_fine > 0 else "传感器1滞后于传感器2"

    return {
        "估计时间偏差 Δt_fine": f"{delay_fine:+.4f} s",
        "时间偏差方向": direction,
        "10Hz融合轨迹采样点数": str(len(t_grid)),
        "融合轨迹时间跨度": f"[{t_grid[0]:.2f} s, {t_grid[-1]:.2f} s]",
        "融合权重": "w₁=0.5, w₂=0.5 (等权)",
        "融合方法": "三次样条插值 + 等权平均",
        "输出文件": "Problem1_10Hz.xlsx",
    }


def _extract_problem2_data() -> Dict[str, str]:
    """从 Problem2_10Hz.xlsx / ablation.xlsx 和 stage2 运行结果提取问题二关键指标。"""
    from core.kalman_filters import estimate_ar1_params, estimate_adaptive_R
    from core.robust_stats import bias_significance_test, compare_bias_methods
    from core.time_alignment import align_sensors
    from core.wavelet_utils import compare_denoise_configs, denoise_trajectory

    file_path = data_path.path2
    df1 = pd.read_excel(file_path, sheet_name="方式1(4Hz)", engine="openpyxl")
    df2 = pd.read_excel(file_path, sheet_name="方式2(5Hz)", engine="openpyxl")

    col_map = {"时间(s)": "t", "X坐标(m)": "x", "Y坐标(m)": "y"}
    df1 = df1.rename(columns=col_map)[["t", "x", "y"]].apply(pd.to_numeric, errors="coerce").dropna()
    df2 = df2.rename(columns=col_map)[["t", "x", "y"]].apply(pd.to_numeric, errors="coerce").dropna()

    t1 = df1["t"].values.astype(np.float64)
    x1 = df1["x"].values.astype(np.float64)
    y1 = df1["y"].values.astype(np.float64)
    t2 = df2["t"].values.astype(np.float64)
    x2 = df2["x"].values.astype(np.float64)
    y2 = df2["y"].values.astype(np.float64)

    # 去噪
    wavelet_opts = ["db4", "sym5"]
    thresh_opts = ["universal", "bayes"]
    r1 = compare_denoise_configs(x1, y1, wavelet_opts, thresh_opts)
    r2 = compare_denoise_configs(x2, y2, wavelet_opts, thresh_opts)
    best1 = min(r1, key=lambda k: r1[k]["accel_var_x"] + r1[k]["accel_var_y"])
    best2 = min(r2, key=lambda k: r2[k]["accel_var_x"] + r2[k]["accel_var_y"])

    x1_d, y1_d = denoise_trajectory(x1, y1, wavelet=best1[0], threshold_method=best1[1])
    x2_d, y2_d = denoise_trajectory(x2, y2, wavelet=best2[0], threshold_method=best2[1])

    # 时间对齐
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        delay, _, _, _, _, _ = align_sensors(
            t1, x1_d, y1_d, t2, x2_d, y2_d,
            target_freq=time_config.target_freq,
            delay_range=alignment_config.delay_range,
            method=alignment_config.method, w1=0.5, w2=0.5,
        )

    # 系统偏差
    t2c = t2 - delay
    sta = max(t1.min(), t2c.min())
    end = min(t1.max(), t2c.max())
    dt = 1.0 / time_config.target_freq
    t_align = sta + np.arange(int((end - sta) / dt) + 1) * dt
    x1a = np.interp(t_align, t1, x1_d)
    y1a = np.interp(t_align, t1, y1_d)
    x2a = np.interp(t_align, t2c, x2_d)
    y2a = np.interp(t_align, t2c, y2_d)

    bias_cmp = compare_bias_methods(x2a, y2a, x1a, y1a)
    bias_x, bias_y = bias_cmp["median"]
    _, _, dx, dy = bias_cmp["median"], bias_cmp["median"][1], x2a - x1a - bias_x, y2a - y1a - bias_y
    # recompute properly
    dx = x2a - x1a - bias_x
    dy = y2a - y1a - bias_y

    # 显著性
    sig_x, p_x = bias_significance_test(dx)
    sig_y, p_y = bias_significance_test(dy)

    # AR(1)
    ar1_alpha, ar1_bias_var = estimate_ar1_params(dx, dy, dt_ref=0.1)
    ar1_rho = np.exp(-ar1_alpha * 0.1)

    # 计算完整方案融合RMSE（避免ablation.xlsx被stage3覆盖）
    from core.kalman_filters import estimate_adaptive_R, fuse_sensors

    R1_est, R2_est = estimate_adaptive_R(t1, x1_d, y1_d, dx, dy, bias_x=bias_x, bias_y=bias_y, method="mad")
    t2f = t2 - delay
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        t_g, x_f, y_f, _, _ = fuse_sensors(
            t1, x1_d, y1_d, t2f, x2_d, y2_d,
            target_freq=time_config.target_freq,
            R1_est=R1_est, R2_est=R2_est,
            ar1_alpha=ar1_alpha, ar1_bias_var=ar1_bias_var,
        )
    x_ref = np.interp(t_g, t1, x1_d)
    y_ref = np.interp(t_g, t1, y1_d)
    rmse_final_val = np.sqrt(np.mean((x_f - x_ref) ** 2 + (y_f - y_ref) ** 2))
    rmse_final = f"{rmse_final_val:.4f} m"

    return {
        "估计时间偏差 Δt_fine": f"{delay:+.4f} s",
        "系统偏差 δx (中位数估计)": f"{bias_x:+.4f} m",
        "系统偏差 δy (中位数估计)": f"{bias_y:+.4f} m",
        "Wilcoxon检验p值 (X方向)": f"p={p_x:.4f} ({'显著' if sig_x else '不显著'})",
        "Wilcoxon检验p值 (Y方向)": f"p={p_y:.4f} ({'显著' if sig_y else '不显著'})",
        "最优去噪小波基 (传感器1)": best1[0],
        "最优去噪小波基 (传感器2)": best2[0],
        "最优阈值策略": best1[1],
        "AR(1)自相关系数 ρ": f"{ar1_rho:.4f}",
        "AR(1)均值回复速率 α": f"{ar1_alpha:.4f} /s",
        "AR(1)平稳方差 σ_b²": f"{ar1_bias_var:.4f} m²",
        "最终融合RMSE (完整方案)": rmse_final,
        "输出文件": "Problem2_10Hz.xlsx",
    }


def _extract_problem3_data() -> Dict[str, str]:
    """从 Problem3_10Hz.xlsx 和 stage3 运行结果提取问题三关键指标。"""
    from core.kalman_filters import estimate_ar1_params, estimate_adaptive_R, fuse_sensors
    from core.robust_stats import bias_significance_test, compare_bias_methods
    from core.time_alignment import align_sensors
    from core.wavelet_utils import compare_denoise_configs, denoise_trajectory

    file_path = data_path.path3
    df1 = pd.read_excel(file_path, sheet_name="方式1(4Hz)", engine="openpyxl")
    df2 = pd.read_excel(file_path, sheet_name="方式2(5Hz)", engine="openpyxl")

    col_map = {"时间(s)": "t", "X坐标(m)": "x", "Y坐标(m)": "y"}
    df1 = df1.rename(columns=col_map)[["t", "x", "y"]].apply(pd.to_numeric, errors="coerce").dropna()
    df2 = df2.rename(columns=col_map)[["t", "x", "y"]].apply(pd.to_numeric, errors="coerce").dropna()

    t1 = df1["t"].values.astype(np.float64)
    x1 = df1["x"].values.astype(np.float64)
    y1 = df1["y"].values.astype(np.float64)
    t2 = df2["t"].values.astype(np.float64)
    x2 = df2["x"].values.astype(np.float64)
    y2 = df2["y"].values.astype(np.float64)

    # 去噪
    wavelet_opts = ["db4", "sym5"]
    thresh_opts = ["universal", "bayes"]
    r1 = compare_denoise_configs(x1, y1, wavelet_opts, thresh_opts)
    r2 = compare_denoise_configs(x2, y2, wavelet_opts, thresh_opts)
    best1 = min(r1, key=lambda k: r1[k]["accel_var_x"] + r1[k]["accel_var_y"])
    best2 = min(r2, key=lambda k: r2[k]["accel_var_x"] + r2[k]["accel_var_y"])

    x1_d, y1_d = denoise_trajectory(x1, y1, wavelet=best1[0], threshold_method=best1[1])
    x2_d, y2_d = denoise_trajectory(x2, y2, wavelet=best2[0], threshold_method=best2[1])

    # 粗搜索
    def _mse(d):
        t2s = t2 + d
        st, ed = max(t1.min(), t2s.min()), min(t1.max(), t2s.max())
        if ed - st < 20:
            return np.inf
        tc = np.linspace(st, ed, max(60, int((ed - st) * 2)))
        return np.mean((np.interp(tc, t1, x1_d) - np.interp(tc, t2s, x2_d)) ** 2 +
                       (np.interp(tc, t1, y1_d) - np.interp(tc, t2s, y2_d)) ** 2)

    ds = np.arange(-500, 800, 5)
    cs = np.array([_mse(d) for d in ds])
    coarse_off = float(ds[np.argmin(cs)])

    # 精细对齐
    t2_shifted = t2 + coarse_off
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        fine_delay, _, _, _, _, _ = align_sensors(
            t1, x1_d, y1_d, t2_shifted, x2_d, y2_d,
            target_freq=time_config.target_freq,
        )
    total_delay = fine_delay - coarse_off

    # 对齐网格
    t2c = t2 - total_delay
    sta = max(t1.min(), t2c.min())
    end = min(t1.max(), t2c.max())
    dt = 1.0 / time_config.target_freq
    t_align = sta + np.arange(int((end - sta) / dt) + 1) * dt
    x1a = np.interp(t_align, t1, x1_d)
    y1a = np.interp(t_align, t1, y1_d)
    x2a = np.interp(t_align, t2c, x2_d)
    y2a = np.interp(t_align, t2c, y2_d)

    # 系统偏差
    bias_cmp = compare_bias_methods(x2a, y2a, x1a, y1a)
    bias_x, bias_y = bias_cmp["median"]
    dx = x2a - x1a - bias_x
    dy = y2a - y1a - bias_y

    # 显著性
    sig_x, p_x = bias_significance_test(dx)
    sig_y, p_y = bias_significance_test(dy)

    # AR(1)
    ar1_alpha, ar1_bias_var = estimate_ar1_params(dx, dy, dt_ref=0.1)
    ar1_rho = np.exp(-ar1_alpha * 0.1)

    # 自适应R vs 默认R
    t2f = t2 - total_delay
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        tgd, xfd, yfd, _, _ = fuse_sensors(
            t1, x1_d, y1_d, t2f, x2_d, y2_d,
            target_freq=time_config.target_freq,
            ar1_alpha=ar1_alpha, ar1_bias_var=ar1_bias_var,
        )
        R1_est, R2_est = estimate_adaptive_R(t1, x1_d, y1_d, dx, dy, bias_x=bias_x, bias_y=bias_y, method="mad")
        tga, xfa, yfa, _, _ = fuse_sensors(
            t1, x1_d, y1_d, t2f, x2_d, y2_d,
            target_freq=time_config.target_freq,
            R1_est=R1_est, R2_est=R2_est,
            ar1_alpha=ar1_alpha, ar1_bias_var=ar1_bias_var,
        )

    cl = min(len(tgd), len(tga))
    xri = np.interp(tgd[:cl], t1, x1_d)
    yri = np.interp(tgd[:cl], t1, y1_d)
    rvdef = np.var(xfd[:cl] - xri) + np.var(yfd[:cl] - yri)
    rvadp = np.var(xfa[:cl] - xri) + np.var(yfa[:cl] - yri)
    adaptive_better = rvadp < rvdef

    return {
        "粗时间偏移估计值": f"{coarse_off:+.2f} s (MSE网格搜索)",
        "精化后总时间偏差 Δt_fine": f"{total_delay:+.4f} s",
        "最优去噪小波基 (传感器1)": best1[0],
        "最优去噪小波基 (传感器2)": best2[0],
        "最优阈值策略": best1[1],
        "系统偏差 δx (中位数估计)": f"{bias_x:+.4f} m",
        "系统偏差 δy (中位数估计)": f"{bias_y:+.4f} m",
        "Wilcoxon检验p值 (X方向)": f"p={p_x:.4f} ({'显著' if sig_x else '不显著'})",
        "Wilcoxon检验p值 (Y方向)": f"p={p_y:.4f} ({'显著' if sig_y else '不显著'})",
        "偏差存在性判定": "双重准则均确认显著存在" if (sig_x and sig_y) else (f"X方向{'显著' if sig_x else '不显著'}, Y方向{'显著' if sig_y else '不显著'}"),
        "自适应R vs 默认R": f"{'自适应R更优' if adaptive_better else '默认R更优或持平'} (残差方差 {rvadp:.4f} vs {rvdef:.4f})",
        "AR(1)自相关系数 ρ": f"{ar1_rho:.4f}",
        "AR(1)均值回复速率 α": f"{ar1_alpha:.4f} /s",
        "AR(1)平稳方差 σ_b²": f"{ar1_bias_var:.4f} m²",
        "输出文件": "Problem3_10Hz.xlsx",
    }


def _extract_problem4_data() -> Dict[str, str]:
    """从 result.xlsx / constraint_stats.xlsx 提取问题四关键指标。"""
    result_path = Path(TABLE_DIR) / "result.xlsx"
    stats_path = Path(TABLE_DIR) / "constraint_stats.xlsx"

    n_shoot, n_photo, n_total = 0, 0, 0
    t_first, t_last = 0.0, 0.0
    solver_info = "未知"

    if result_path.exists():
        df = pd.read_excel(result_path, engine="openpyxl")
        n_total = len(df)
        n_shoot = int((df["任务"] == "射击").sum()) if "任务" in df.columns else 0
        n_photo = n_total - n_shoot
        if "任务执行时刻(s)" in df.columns:
            t_first = df["任务执行时刻(s)"].min() - 1.5
            t_last = df["任务执行时刻(s)"].max()

    # 从 constraint_stats 获取覆盖信息
    coverage_str = "需根据实际结果填入"
    if stats_path.exists():
        df_s = pd.read_excel(stats_path, engine="openpyxl")
        uncovered_all = df_s[(df_s["阶段"] == "未覆盖目标") & (df_s["类别"] == "全部")]
        if len(uncovered_all) > 0:
            uncovered_count = int(uncovered_all["数量"].iloc[0])
            # 从附件4读目标总数
            try:
                target_path = Path(data_path.file4)
                df_tgts = pd.read_excel(target_path, engine="openpyxl")
                total_tgts = len(df_tgts)
            except Exception:
                total_tgts = 18  # fallback
            total_cov = total_tgts - uncovered_count
            coverage_str = f"{total_cov} / {total_tgts} ({100 * total_cov / max(total_tgts, 1):.1f}%)"

    # 尝试检测求解器
    try:
        import pulp
        solver_info = "ILP (PuLP + CBC)"
    except ImportError:
        solver_info = "贪心调度 Greedy_v3 (PuLP未安装, ILP回退)"

    return {
        "调度求解器": solver_info,
        "射击任务数": str(n_shoot),
        "拍照任务数": str(n_photo),
        "总任务数": str(n_total),
        "覆盖目标数 / 总目标数": coverage_str,
        "任务时间跨度": f"[{t_first:.2f} s, {t_last:.2f} s]" if n_total > 0 else "N/A",
        "输出文件": "result.xlsx",
    }


# ============================================================
#  表生成入口
# ============================================================

def build_table1() -> List[Tuple[str, str]]:
    """构建表1：问题一关键结果汇总表。"""
    return list(_extract_problem1_data().items())


def build_table2() -> List[Tuple[str, str]]:
    """构建表2：问题二关键结果汇总表。"""
    return list(_extract_problem2_data().items())


def build_table3() -> List[Tuple[str, str]]:
    """构建表3：问题三关键结果汇总表。"""
    return list(_extract_problem3_data().items())


def build_table4() -> List[Tuple[str, str]]:
    """构建表4：问题四任务调度结果表。"""
    return list(_extract_problem4_data().items())


def build_table5() -> Tuple[List[str], List[List[str]]]:
    """构建表5：跨问题指标对比表。

    Returns
    -------
    headers : list
        列标题 ["指标", "问题一", "问题二", "问题三"]
    rows : list of list
        每行数据
    """
    p1 = _extract_problem1_data()
    p2 = _extract_problem2_data()
    p3 = _extract_problem3_data()

    # 问题二的偏差量级
    bias_x2 = float(p2["系统偏差 δx (中位数估计)"].split()[0])
    bias_y2 = float(p2["系统偏差 δy (中位数估计)"].split()[0])
    # 问题三的偏差量级
    bias_x3 = float(p3["系统偏差 δx (中位数估计)"].split()[0])
    bias_y3 = float(p3["系统偏差 δy (中位数估计)"].split()[0])

    # p值
    get_p = lambda d, key: d.get(key, "").split("=")[1].split(" ")[0] if "p=" in d.get(key, "") else "N/A"
    p2x = get_p(p2, "Wilcoxon检验p值 (X方向)")
    p2y = get_p(p2, "Wilcoxon检验p值 (Y方向)")
    p3x = get_p(p3, "Wilcoxon检验p值 (X方向)")
    p3y = get_p(p3, "Wilcoxon检验p值 (Y方向)")

    # RMSE
    rmse2 = p2.get("最终融合RMSE (完整方案)", "N/A")

    headers = ["指标", "问题一", "问题二", "问题三"]
    rows = [
        ["数据特性", "无噪声", "含噪声+系统偏差", "实际测量数据"],
        ["时间偏差 Δt", p1["估计时间偏差 Δt_fine"], p2["估计时间偏差 Δt_fine"], p3["精化后总时间偏差 Δt_fine"]],
        ["系统偏差量级 (X/Y)", "不存在", f"{bias_x2:+.2f} / {bias_y2:+.2f} m", f"{bias_x3:+.2f} / {bias_y3:+.2f} m"],
        ["Wilcoxon p值 (X)", "不适用", f"p={p2x}", f"p={p3x}"],
        ["Wilcoxon p值 (Y)", "不适用", f"p={p2y}", f"p={p3y}"],
        ["偏差存在性", "无",
         "显著存在" if ("显著" in p2.get("Wilcoxon检验p值 (X方向)", "") or "显著" in p2.get("Wilcoxon检验p值 (Y方向)", "")) else "待定",
         "X显著/Y不显著" if ("显著" in p3.get("Wilcoxon检验p值 (X方向)", "") and "不显著" in p3.get("Wilcoxon检验p值 (Y方向)", ""))
         else ("显著存在" if ("显著" in p3.get("Wilcoxon检验p值 (X方向)", "") or "显著" in p3.get("Wilcoxon检验p值 (Y方向)", "")) else "待定")],
        ["融合RMSE", "理论最优", rmse2, "因无真值不可直接计算"],
        ["融合方法", "等权平均", "AR(1)+EKF+自适应R", "AR(1)+EKF+自适应R"],
        ["去噪方法", "无", f"{p2['最优去噪小波基 (传感器1)']} + {p2['最优阈值策略']}",
         f"{p3['最优去噪小波基 (传感器1)']} + {p3['最优阈值策略']}"],
        ["输出频率", "10 Hz", "10 Hz", "10 Hz"],
        ["融合轨迹点数", p1["10Hz融合轨迹采样点数"],
         str(len(pd.read_excel(Path(TABLE_DIR) / "Problem2_10Hz.xlsx", engine="openpyxl"))),
         str(len(pd.read_excel(Path(TABLE_DIR) / "Problem3_10Hz.xlsx", engine="openpyxl")))],
        ["时间跨度", p1["融合轨迹时间跨度"],
         _get_timespan(Path(TABLE_DIR) / "Problem2_10Hz.xlsx"),
         _get_timespan(Path(TABLE_DIR) / "Problem3_10Hz.xlsx")],
    ]
    return headers, rows


def _get_timespan(path: Path) -> str:
    """读取 xlsx 的时间跨度字符串。"""
    if not path.exists():
        return "N/A"
    df = pd.read_excel(path, engine="openpyxl")
    t_col = "Time(s)" if "Time(s)" in df.columns else df.columns[0]
    return f"[{df[t_col].min():.2f}, {df[t_col].max():.2f}] s"


# ============================================================
#  Excel 写入
# ============================================================

def generate_all_tables(output_path: Optional[Path] = None) -> Path:
    """生成全部五个汇总表并写入 Excel 文件。

    Parameters
    ----------
    output_path : Path, optional
        输出路径，默认为 output/tables/Summary_Tables.xlsx

    Returns
    -------
    Path
        实际输出文件路径
    """
    if output_path is None:
        output_path = Path(TABLE_DIR) / "Summary_Tables.xlsx"

    if not _HAS_OPENPYXL:
        raise ImportError("需要安装 openpyxl: pip install openpyxl")

    ensure_dirs()
    from openpyxl import Workbook

    wb = Workbook()
    s = _default_styles()

    # ── 表1 ──
    ws1 = wb.active
    ws1.title = "表1-问题一汇总"
    _write_two_col(ws1, "表1：问题一关键结果汇总表", build_table1())
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 32

    # ── 表2 ──
    ws2 = wb.create_sheet("表2-问题二汇总")
    _write_two_col(ws2, "表2：问题二关键结果汇总表", build_table2())
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 32

    # ── 表3 ──
    ws3 = wb.create_sheet("表3-问题三汇总")
    _write_two_col(ws3, "表3：问题三关键结果汇总表", build_table3())
    ws3.column_dimensions["A"].width = 38
    ws3.column_dimensions["B"].width = 42

    # ── 表4 ──
    ws4 = wb.create_sheet("表4-问题四汇总")
    _write_two_col(ws4, "表4：问题四任务调度结果表", build_table4())
    ws4.column_dimensions["A"].width = 32
    ws4.column_dimensions["B"].width = 50

    # ── 表5 ──
    ws5 = wb.create_sheet("表5-跨问题对比")
    headers5, rows5 = build_table5()
    ws5.merge_cells("A1:D1")
    ws5.cell(row=1, column=1, value="表5：跨问题指标对比表").font = s["title_font"]
    ws5.cell(row=1, column=1).alignment = s["center_align"]
    for j, h in enumerate(headers5, 1):
        ws5.cell(row=2, column=j, value=h)
    for i, row in enumerate(rows5, 3):
        for j, val in enumerate(row, 1):
            ws5.cell(row=i, column=j, value=val)
    _style_multi_col_sheet(ws5, 2, len(rows5), 4)
    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 22
    ws5.column_dimensions["C"].width = 28
    ws5.column_dimensions["D"].width = 28

    wb.save(output_path)
    print(f"[generate_summary] 汇总表已保存: {output_path} ({output_path.stat().st_size / 1024:.1f} KB)")
    return output_path


def _write_two_col(ws, title: str, data: List[Tuple[str, str]]):
    """写入双列表格（指标 + 数值）。"""
    s = _default_styles()
    ws.merge_cells("A1:B1")
    c = ws.cell(row=1, column=1, value=title)
    if s:
        c.font = s["title_font"]
        c.alignment = s["center_align"]

    ws.cell(row=2, column=1, value="指标")
    ws.cell(row=2, column=2, value="数值")
    for i, (k, v) in enumerate(data, 3):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

    _style_two_col_sheet(ws, 2, len(data))


# ============================================================
#  主入口
# ============================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="生成四个问题的关键结果汇总表")
    parser.add_argument("--output", "-o", type=str, default=None,
                       help="输出 Excel 路径 (默认: output/tables/Summary_Tables.xlsx)")
    args = parser.parse_args()

    out = Path(args.output) if args.output else None
    generate_all_tables(out)
    print("[generate_summary] 完成。")
